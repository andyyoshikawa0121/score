import glob
import subprocess
import os
import pathlib
import shutil
import openpyxl


def rename_files():
    # ファイルのrename
    files_dir = "files"
    files_extension = ".c"

    files = glob.glob(f"./{files_dir}/*{files_extension}")

    student_num_list = list(
        map(lambda file: file.split()[0].replace("./{files_dir}/", ""), files)
    )
    for index, file in enumerate(files):
        os.rename(file, f"{student_num_list[index]}{files_extension}")


def calc_score():
    files_dir = "files"
    files_extension = ".c"
    src_dir = "src"
    excel_file = "saiten.xlsx"
    excel_sheet_name = "Sheet2"
    gcc_out_file = "./a.exe"
    files = glob.glob(f"./{files_dir}/*{files_extension}")
    input_files = glob.glob(f"./{src_dir}/*.txt")
    wb = openpyxl.load_workbook(excel_file)
    ws = wb[excel_sheet_name]

    for student_index, file in enumerate(files):
        # 学籍番号をシートに書き込み
        cell = ws.cell(row=student_index + 1, column=1)
        cell.value = file.replace("\\", "/").replace(f"./{files_dir}/", "").replace(files_extension, "")

        for input_index, input in enumerate(input_files):
            shutil.copy(input, "./input.txt")
            os.system("gcc %s" % file)
            # 正しくコンパイルされ a.outがあるか
            if os.path.exists(gcc_out_file):
                result = (
                    subprocess.Popen(gcc_out_file, stdout=subprocess.PIPE)
                    .communicate()[0]
                    .decode("utf-8")
                )

                # 出力をシートに書き込み
                output_cell = ws.cell(row=student_index + 1, column=input_index + 2)
                output_cell.value = result
                # a.outを削除
                os.remove(gcc_out_file)
            else:
                # 出力をシートに書き込み
                output_cell = ws.cell(row=student_index + 1, column=input_index + 2)
                output_cell.value = "compile was failed"

    wb.save(excel_file)
    wb.close()


if __name__ == "__main__":
    # moodle からダウンロードしたファイルをrenameしたいとき
    rename_files()


    # ファイルを実行しexcelファイルに記入したいとき
    # calc_score()
